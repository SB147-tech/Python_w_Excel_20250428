import openpyxl

wb = openpyxl.load_workbook("videogamesales2.xlsx")
ws = wb.active

ws = wb['vgsales']

ws['P1']= 'Average Sales'
ws["P2"] = '=AVERAGE(K2:K16220)'

ws['Q1'] = 'Number of populated cells'
ws['Q2'] = '=COUNTA(E2:E16220)'

ws['R1'] = 'Number of rows with Sports Genre'
ws['R2'] = '=COUNTIF(E2:E16220,"Sports")'

ws['S1'] = 'Total sport sales'
ws['S2'] = '=SUMIF(E2:E16220,"Sports",K2:K16220)'

ws['T1'] = 'Rounded Sum of Sport Sales'
ws['T2'] = '=CEILING(S2,25)'

wb.save('videogamesales2.xlsx')
wb.close()