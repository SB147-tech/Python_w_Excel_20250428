import openpyxl

wb = openpyxl.load_workbook("videogamesales2.xlsx")
ws = wb.active

ws = wb['vgsales']

# nowy wiersz
new_row =(1, 'Duaaaaaa', 1986, 'Action', 'Nintendo', 3.74, 0.93, 1.69, 0.14, 6.51, 6.5)
new_row2 =(1, 'Duaaaaaa2', 1986, 'Action', 'Nintendo', 3.74, 0.93, 1.69, 0.14, 6.51, 6.5)
ws.append(new_row)
ws.append(new_row2)

wb.save('videogamesales2.xlsx')
wb.close()

# sprawdzanie czy zapisal

# odczytanie ostatniego wiersza
values = [ws.cell(row=ws.max_row, column=i).value for i in range(1, ws.max_column + 1)]
values2 = [ws.cell(row=ws.max_row, column=i).value for i in range(1, ws.max_column + 1)]
print(values)
print(values2)

# ususniecie jednego ostatniego wiersza
ws.delete_rows(ws.max_row, 1)

wb.save('videogamesales2.xlsx')
wb.close()