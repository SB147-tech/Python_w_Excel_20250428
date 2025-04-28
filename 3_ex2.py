from openpyxl import Workbook, load_workbook

wb = Workbook()
ws = wb.active

ws['A1'] = 42

ws.append([1, 2, 3])

wb.save("sample.xlsx")
wb.close()

print("Excel file was created successfully.")
print("file nameL sample.xlsx")
print("Location in current directory.")

workbook = load_workbook("sample.xlsx")
sheet = workbook.active
print(sheet)

print(sheet['A1'].value)

for row in sheet.iter_rows(min_row=1, max_row=3):
    for cell in row:
        print(cell.value)