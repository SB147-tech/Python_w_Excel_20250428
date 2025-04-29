import openpyxl
from openpyxl.formatting.rule import CellIsRule

from openpyxl.styles import Font, colors, PatternFill, Border, Side

wb = openpyxl.load_workbook("videogamesales2.xlsx")
ws = wb.active

ws = wb['Video Games Sales Data']

ws['A1'].font = Font(color='FF0000', bold=True, size=12)
ws['A2'].font = Font(color='0000FF')

fill = PatternFill(
    start_color='90EE90',
    end_color='90EE90', fill_type='solid'
)

ws['A1'].fill = PatternFill('lightVertical', start_color='38e3ff')

my_border = Side(border_style='thin', color='000000')
ws['A1'].border = Border(
    top=my_border, left=my_border, right=my_border, bottom=my_border
)

ws.conditional_formatting.add(
    'G2:K16594',CellIsRule(operator='greaterThan', formula=[8], fill=fill, font=Font(color='FF0000'))
)

wb.save('videogamesales2.xlsx')
wb.close()